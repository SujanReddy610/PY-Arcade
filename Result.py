#importing openpyxl and creating a sheet via work book
import openpyxl
import datetime
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active

row=sheet.max_row
col=sheet.max_column

#deleting the result row

s=["Dots and Boxes","Hangman","Madlibs","Stone, Paper and Scissors","Tic Tac Toe","Number Guessing","Guess the colour","Escape the Room"]


improvise=sheet.cell(row=1,column=col)
if improvise.value=="Result":
    x=0
    ls=[]
    for i in range(2,col):
        y=sheet.cell(row=3,column=i)
        if y.value==1:
            x+=1
    for j in range(3,col):
        y=sheet.cell(row=1,column=j)
        ls.append(y.value)
    
    if x==8:
        final_1=sheet.cell(row=4,column=col)
        final_2=sheet.cell(row=5,column=col)
        print("The final score of player 1 is ",final_1.value)
        print("The final score of player 2 is ",final_2.value)
        if final_1.value<final_2.value:
            print("Player 2 is the winner of the tournament")
        elif final_1.value>final_2.value:
            print("Player 1 is the winner of the tournament")
        else:
            print("It's a tie between the players")
    else:
        final_1=sheet.cell(row=4,column=col)
        final_2=sheet.cell(row=5,column=col)
        print("The score of player 1 is ",final_1.value)
        print("The score of player 2 is ",final_2.value)
        if final_1.value<final_2.value:
            print("Player 2 is in the lead for now")
        elif final_1.value>final_2.value:
            print("Player 1 is in the lead for now")
        else:
            print("It's a tie between the players for now")
        for i in s:
            if i not in ls:
                print("{0} has to be played".format(i))
        
        
    


    
else:
    row=sheet.max_row
    col=sheet.max_column

    change_1=sheet.cell(row=1,column=col+1)
    change_1.value="Result"
    x=0
    fscores_1=0
    fscores_2=0
    x=0
    ls=[]
    for j in range(3,col+2):
        y=sheet.cell(row=1,column=j)
        ls.append(y.value)
    
    for i in range(3,col+1):
        score_1=sheet.cell(row=4,column=i)
        score_2=sheet.cell(row=5,column=i)
        fscores_1+=score_1.value
        fscores_2+=score_2.value
    change_2=sheet.cell(row=4,column=col+1)
    change_2.value=fscores_1
    change_3=sheet.cell(row=5,column=col+1)
    change_3.value=fscores_2

    x=0
    for i in range(2,col):
        y=sheet.cell(row=3,column=i)
        if y.value==1:
            x+=1

    if x==8:
        final_1=sheet.cell(row=4,column=col+1)
        final_2=sheet.cell(row=5,column=col+1)
        print("The final score of player 1 is ",final_1.value)
        print("The final score of player 2 is ",final_2.value)
        if final_1.value<final_2.value:
            print("Player 2 is the winner of the tournament")
        elif final_1.value>final_2.value:
            print("Player 1 is the winner of the tournament")
        else:
            print("It's a tie between the players")
    else:
        final_1=sheet.cell(row=4,column=col+1)
        final_2=sheet.cell(row=5,column=col+1)
        print("The score of player 1 is ",final_1.value)
        print("The score of player 2 is ",final_2.value)
        if final_1.value<final_2.value:
            print("Player 2 is in the lead for now")
        elif final_1.value>final_2.value:
            print("Player 1 is in the lead for now")
        else:
            print("It's a tie between the players for now")
        for i in s:
            if i not in ls:
                print("{0} has to be played".format(i))


 




        
        
    

    wb.save('final.xlsx')

