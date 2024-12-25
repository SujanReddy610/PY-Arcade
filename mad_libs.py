import random
import tkinter

#importing openpyxl and creating a sheet via work book
import openpyxl
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active
row=sheet.max_row
col=sheet.max_column

def get_noun(number):
    noun = []
    for i in range(number):
        n = input('enter a noun: ')
        noun.append(n.upper())
    return noun

def get_plural_noun(number):
    plural_noun = []
    for i in range(number):
        n = input("enter a plural noun: ")
        plural_noun.append(n.upper())
    return plural_noun

def get_adjective(number):
    adjective= []
    for i in range(number):
        n = input("enter an adjective: ")
        adjective.append(n.upper())
    return adjective

def get_verb(number):
    verb = []
    for i in range(number):
        n = input("enter a verb(present tense): ")
        verb.append(n.upper())
    return verb

def get_body_part(number):
    part_of_the_body = []
    for i in range(number):
        n = input("enter a part of the body(plural): ")
        part_of_the_body.append(n.upper())
    return part_of_the_body

class story1:
    def __init__(self):
        print()
        self.noun = get_noun(1)
        self.plural_noun = get_plural_noun(2)
        self.adjective = get_adjective(2)
        self.verb = get_verb(2)
        self.body_part = get_body_part(1)
    def print_story(self):
        print()
        print("Today, every student has a computer small enough to fit into his " + self.noun[0] + '.')
        print("He can solve any math problem by simply pushing the computer's little " + self.plural_noun[0] + '.')
        print("Computers can add, multiply, divide and " + self.verb[0] + '.')
        print("They can also " + self.verb[1] + " better than a human.")
        print("Some computers have " + self.body_part[0] + '.')
        print("Others have a/an " + self.adjective[0] + " screen that shows all kind of " + self.plural_noun[1] + " and " + self.adjective[1] + " figures.")

class story2:
    def __init__(self):
        print()
        self.adjective = get_adjective(5)
        self.verb = get_verb(2)
        self.noun = get_noun(2)
    def print_story(self):
        print()
        print("That restaurant has really " + self.adjective[0] + " food.")
        print("Just thinking about it makes my stomach " + self.verb[0] + '.')
        print("The sphagetti is " + self.adjective[1] + " and tastes like " + self.noun[0] + '.')
        print("One day, I swear one of my meatballs started to " + self.verb[1] + '!')
        print("The chicken tacos are totally " + self.adjective[2] + " and look kind of like old " + self.noun[1] + '.')
        print("My friend Ravi actually likes the Dosas, even though it's " + self.adjective[3] + " and " + self.adjective[4] + '.')

def start_game():
    a = [1, 2]
    choice = random.choice(a)
    if choice == 1:
        print("FIRST TURN - PLAYER 1")
        person1 = story1()

        print("SECOND TURN - PLAYER 2")
        person2 = story1()
    
        print("player 1's story - ")
        person1.print_story()
        print("player 2's story - ")
        person2.print_story()
    else:
        print("FIRST TURN - PLAYER 1")
        person1 = story2()

        print("SECOND TURN - PLAYER 2")
        person2 = story2()

        print("player 1's story - ")
        person1.print_story()
        print("player 2's story - ")
        person2.print_story()

    print()
    print('Please vote for the funnier/sillier story: Player1 or Player2, upon mutual understanding')
    winner = input("Type: Player1 or Player2 or Both\n")

    #________EXCEL Part starts___________
    row=sheet.max_row
    col=sheet.max_column
    #deleting the result row
    improvise=sheet.cell(row=1,column=col)
    if improvise.value=="Result":
        improvise.value=None
        sheet.delete_cols(col)
        row=sheet.max_row
        col=sheet.max_column
        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Madlibs"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Multiplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
       #   appending scores
        if winner=="player1" or winner=="Player1":
            c1.value=10
            c2.value=0
        elif winner=="player2" or winner=="Player2":
            c1.value=0
            c2.value=10
        else:
            c1.value=5
            c2.value=5 
    

    else:
        row=sheet.max_row
        col=sheet.max_column

        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Madlibs"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Multiplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #   appending scores
        if winner=="player1" or winner=="Player1":
            c1.value=10
            c2.value=0
        elif winner=="player2" or winner=="Player2":
            c1.value=0
            c2.value=10
        else:
            c1.value=5
            c2.value=5

    x=1
    for i in range(1,col+1):
        y=sheet.cell(row=1,column=i)
        if y.value=="Madlibs":
            z=sheet.cell(row=3,column=i)
            x=z.value+1
        else:
            continue

    change_3=sheet.cell(row=3,column=col+1)
    change_3.value=x 
    

    wb.save('final.xlsx')


    return

'''root=tkinter.Tk()
root.geometry('10x10')
root.title("PY-ARCADE")
start_game()'''
