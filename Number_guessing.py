import random
import math

#importing openpyxl and creating a sheet via work book
import openpyxl
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active
row=sheet.max_row
col=sheet.max_column



def start_game():
    # Taking Inputs
    lower = int(input("Enter Lower bound:- "))

    # Taking Inputs
    upper = int(input("Enter Upper bound:- "))

    # generating random number between
    # the lower and upper
    x = random.randint(lower, upper)
    if (upper-lower)<=45:
        print("\n\tYou've only ",
              round(math.log(upper - lower + 1, 2)+1),
              " chances to guess the integer!\n")
    else:
        print("\n\tYou've only ",
              round(math.log(upper - lower + 1, 2)),
              " chances to guess the integer!\n")
    # Initializing the number of guesses and score.
    count = 0
    score=0

    # for calculation of minimum number of
    # guesses depends upon range
    while count < math.log(upper - lower + 1, 2):
        count += 1

        # taking guessing number as input
        guess = int(input("Guess a number:- "))

        # Condition testing
        if x == guess:
            print("Congratulations you did it in ",
                  count, " try")
            # Once guessed, loop will break
            if count==1:
                score=10
                print("Your score is: ",score)
            else:
                score=5
                print("Your score is: ",score)
            break
        elif x > guess:
            print("You guessed too small!")
        elif x < guess:
            print("You Guessed too high!")

    # If Guessing is more than required guesses,
    # shows this output.
    if count >= math.log(upper - lower + 1, 2):
        print("\nThe number is %d" % x)
        print("\tBetter Luck Next time!")
        print("Your score is: ",score)
    
    #________EXCEL Part starts___________
    row=sheet.max_row
    col=sheet.max_column
    #deleting the result row
    improvise=sheet.cell(row=1,column=col)
    if improvise.value=="Result":
        improvise.value=None
        sheet.delete_cols(col)
        row=sheet.max_row
        col=sheet.max_column
        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Number Guessing"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #appending scores
        if score==10:
            c1.value=10
            c2.value=0
        elif score==5:
            c1.value=5
            c2.value=5
        else:
            c1.value=0
            c2.value=10
    

    else:
        row=sheet.max_row
        col=sheet.max_column

        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Number Guessing"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"

        x=0
        
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #   appending scores
        if score==10:
            c1.value=10
            c2.value=0
        elif score==5:
            c1.value=5
            c2.value=5
        else:
            c1.value=0
            c2.value=10
    x=1
    for i in range(1,col+1):
        y=sheet.cell(row=1,column=i)
        if y.value=="Number Guessing":
            z=sheet.cell(row=3,column=i)
            x=z.value+1
        else:
            continue

    change_3=sheet.cell(row=3,column=col+1)
    change_3.value=x
    

    wb.save('final.xlsx')



#start_game()
