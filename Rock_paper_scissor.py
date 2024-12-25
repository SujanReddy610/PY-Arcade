import random
comp_wins=0
player_wins=0

#importing openpyxl and creating a sheet via work book
import openpyxl
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active
row=sheet.max_row
col=sheet.max_column

def choose_option():
    user_choice=input("choose Rock, Paper or Scissors: ")
    if user_choice in ['Rock', 'rock', 'R', 'r']:
        user_choice='r'
    elif user_choice in ['Paper', 'paper', 'P', 'p']:
        user_choice='p'
    elif user_choice in ['Scissors', 'scissors', 'S', 's']:
        user_choice='s'
    else:
        print("I dont understand, try again")
        choose_option()
    return user_choice
def computer_option():
    comp_choice=random.randint(1,3)
    if comp_choice==1:
        comp_choice='r'
    elif comp_choice==2:
        comp_choice='p'
    else:
        comp_choice='s'
    return comp_choice
def start_game():
    global player_wins,comp_wins
    for i in range(0,3):
        
        print("")
        user_choice=choose_option()
        comp_choice=computer_option()
        print("")
        if user_choice=='r':
            if comp_choice=='r':
                print("you chose rock. computer chose rock. you tied")
            elif comp_choice=='p':
                print("you chose rock. computer chose paper. you lose")
                comp_wins+=1
            elif comp_choice=='s':
                
                print("you chose rock. computer chose scissors. you won")
                player_wins+=1
        elif user_choice=='p':
            if comp_choice=='r':
                print("you chose paper. computer chose rock. you won")
                player_wins+=1
            elif comp_choice=='p':
                print("you chose paper. computer chose paper. you tied")
            elif comp_choice=='s':
                print("you chose paper. computer chose scissors. you lose")
                comp_wins+=1
        elif user_choice=='s':
        
            if comp_choice=='r':
                print("you chose scissors. computer chose rock. you lose")
                comp_wins+=1
            elif comp_choice=='p':
                print("you chose scissors. computer chose paper. you won")
                player_wins+=1
            elif comp_choice=='s':
                print('you chose scissors. computer chose scissors. you tied')
        print("")
        print("player wins: " + str(player_wins))
        print("computer wins: " + str(comp_wins))
        print("")
    player_score= player_wins
    computer_score= comp_wins
    print("    *********     ")
    print("Your final score = ",player_score)
    print("Computer final score = ",computer_score)
    if(player_score<computer_score):
        print("Computer won the game")
    elif(player_score==computer_score):
        print("The game is tied between you and computer")
    else:
        print("You won the game")

   #________EXCEL Part starts___________
    row=sheet.max_row
    col=sheet.max_column
    #deleting the result row
    improvise=sheet.cell(row=1,column=col)
    if improvise.value=="Result":
        improvise.value=None
        sheet.delete_cols(col)
        row=sheet.max_row
        col=sheet.max_column
        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Stone, Paper and Scissors"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #appending scores
        if player_score>computer_score:
            c1.value=10
            c2.value=0
        elif player_score<computer_score:
            c1.value=0
            c2.value=10
        else:
            c1.value=5
            c2.value=5
    

    else:
        row=sheet.max_row
        col=sheet.max_column

        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Stone, Paper and Scissors"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #appending scores
        if player_score>computer_score:
            c1.value=10
            c2.value=0
        elif player_score<computer_score:
            c1.value=0
            c2.value=10
        else:
            c1.value=5
            c2.value=5

    x=1
    for i in range(1,col+1):
        y=sheet.cell(row=1,column=i)
        if y.value=="Stone, Paper and Scissors":
            z=sheet.cell(row=3,column=i)
            x=z.value+1
        else:
            continue

    change_3=sheet.cell(row=3,column=col+1)
    change_3.value=x  
    

    wb.save('final.xlsx')

 

#start_game()

   
            
            
            

                
        
        
    
