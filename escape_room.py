import random

#importing openpyxl and creating a sheet via work book
import openpyxl
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active
row=sheet.max_row
col=sheet.max_column

player = 100
enemy =100
def enemyattack():
    global player
    enemyattack=random.randint(4,15)
    print("Enemy attacks!")
    print("Attack =",enemyattack,"HP")
    player=player-enemyattack
    print("player health :",player)
def conserveattack():
    global enemy
    attack=random.randint(7,9)
    print("player attacks!")
    print("Attack =",attack,"HP")
    enemy=enemy-attack
    print("Enemy Health : ",enemy)
def hardattack():
    global enemy
    attack=random.randint(1,17)
    print("player attacks!")
    print("Attack = ",attack,"HP")
    enemy=enemy-attack
    print("Enemy health :",enemy)
'''def checkplayerhealth():
    import sys
    if player < 1:
        print("You died")
        sys.exit()'''

def start_game():
    global enemy, player
    print("You wake up in a dark room. You dont know where you are and how you got there")
    print("You see two doors in the room ")
    while True:
        print("do you want to go through one of the door")
        print("1. left door")
        print("2. right door")
        print("3. just sit here a minute")
        answer=input("select one among the three options")
        if answer=='1':
            print("good choice, but be prepared for a fight.")
            break
        elif answer=='2':
            print("You like hard way, don't you?")
            enemy=enemy*2
            print("enemy=",enemy)
            break
        elif answer=='3':
            print("You are wasting time! You're losing your life")
            player=player-10
            print("player Health : ",player)
        else:
            print("That's not an option, try something else")
    print("You can hear the sounds of a monster approaching in the darkness. what do you want to do?")
    while True:
        if player < 1:
            print("You died")
            break
        if enemy<1:
            print("You defeated the monster")
            break
        print("Fight or Flee")
        choice=input()
        if choice in ['flee','Flee']:
            print("You gave up, you lost")
            break
        elif choice in ['fight','Fight']:
            print("choose 1 for hardattack. choose 2 for conserveattack")
            choice=input()
            if choice=='1':
                hardattack()
            elif choice=='2':
                conserveattack()
            else:
                print("That's not an option. Try again.")
        else:
            print("That's not an option. Try again.")
        print()
        enemyattack()
        print()
    #________EXCEL Part starts___________
    row=sheet.max_row
    col=sheet.max_column
    #deleting the result row
    improvise=sheet.cell(row=1,column=col)
    if improvise.value=="Result":
        improvise.value=None
        sheet.delete_cols(col)
        row=sheet.max_row
        col=sheet.max_column
        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Escape the Room"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #appending scores
        if enemy<1:
            c1.value=10
            c2.value=0
        elif player<1:
            c1.value=0
            c2.value=10
        else:
            c1.value=5
            c2.value=5
        

    else:
        row=sheet.max_row
        col=sheet.max_column

        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Escape the Room"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #appending scores
        if enemy<1:
            c1.value=10
            c2.value=0
        elif player<1:
            c1.value=0
            c2.value=10
        else:
            c1.value=5
            c2.value=5
    x=1
    for i in range(1,col+1):
        y=sheet.cell(row=1,column=i)
        if str(y.value)=="Escape the Room":
            z=sheet.cell(row=3,column=i)
            x=int(z.value)+1
        else:
            continue

    change_3=sheet.cell(row=3,column=col+1)
    change_3.value=x 
    

    wb.save('final.xlsx')


#start_game()

        
