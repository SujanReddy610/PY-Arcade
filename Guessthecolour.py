# import the modules 
import tkinter
import random
from tkinter import messagebox

#importing openpyxl and creating a sheet via work book
import openpyxl
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active
row=sheet.max_row
col=sheet.max_column

# list of possible colour.
colours = ['Red','Blue','Green','Pink','Black',
           'Yellow','Orange','White','Purple','Brown']
score = 0
  
# the game time left, initially 30 seconds.
timeleft = 30
  
# function that will start the game.
def startGame(event):
      
    if timeleft == 30:
          
        # start the countdown timer.
        countdown()
          
    # run the function to
    # choose the next colour.
    nextColour()
  
# Function to choose and
# display the next colour.
def nextColour():
  
    # use the globally declared 'score'
    # and 'play' variables above.
    global score
    global timeleft
  
    # if a game is currently in play
    if timeleft > 0:
  
        # make the text entry box active.
        e.focus_set()
  
        # if the colour typed is equal
        # to the colour of the text
        if e.get().lower() == colours[1].lower():
              
            score += 1
  
        # clear the text entry box.
        e.delete(0, tkinter.END)
          
        random.shuffle(colours)
          
        # change the colour to type, by changing the
        # text _and_ the colour to a random colour value
        label.config(fg = str(colours[1]), text = str(colours[0]))
          
        # update the score.
        scoreLabel.config(text = "Score: " + str(score))
        
    else:
        S1="Your Score ="+str(score)
        if score>=8:
            box=messagebox.showinfo("Congracts",S1)
        elif score>=5:
            box=messagebox.showinfo("Well done",S1)
        else:
            box=messagebox.showinfo("Good Try",S1)
        
        
  
# Countdown timer function 
def countdown():
  
    global timeleft
  
    # if a game is in play
    if timeleft > 0:
  
        # decrement the timer.
        timeleft -= 1
          
        # update the time left label
        timeLabel.config(text = "Time left: "
                               + str(timeleft))
                                 
        # run the function again after 1 second.
        timeLabel.after(1000, countdown)
  
while(True):
    score=0
    timeleft=30
    # Driver Code
  
    # create a GUI window
    root = tkinter.Tk()
      
    # set the title
    root.title("COLORGAME")

    # set icon
    root.iconbitmap(r"OIP.ico")

    # set the size
    root.geometry("375x375")

    #set background
    root.configure(background='powder blue')
      
    # add an instructions label
    instructions = tkinter.Label(root, text = "Type in the colour\n"
                            "of the words,not the word text!",
                                          font = ('Helvetica', 12))
    instructions.pack() 
      
    # add a score label
    scoreLabel = tkinter.Label(root, text = "Press enter to start",
                                          font = ('Helvetica', 12))
    scoreLabel.pack()
      
    # add a time left label
    timeLabel = tkinter.Label(root, text = "Time left: " +
                  str(timeleft), font = ('Helvetica', 12))
                    
    timeLabel.pack()
      
    # add a label for displaying the colours
    label = tkinter.Label(root, font = ('Helvetica', 60))
    label.pack()
      
    # add a text entry box for
    # typing in colours
    e = tkinter.Entry(root)
      
    # run the 'startGame' function 
    # when the enter key is pressed
    root.bind('<Return>', startGame)
    e.pack()
      
    # set focus on the entry box
    e.focus_set()
      
    # start the GUI
    root.mainloop()
    if score>=8:
        print("Well done, Your score is {0}, You win".format(score))
    elif score>=5:
        print("Well done, Your score is {0}, It's a tie".format(score))
    else:
        print("Good try, your score is {0}, you lose")

    #________EXCEL Part starts___________
    row=sheet.max_row
    col=sheet.max_column
    #deleting the result row
    improvise=sheet.cell(row=1,column=col)
    if improvise.value=="Result":
        improvise.value=None
        sheet.delete_cols(col)
        row=sheet.max_row
        col=sheet.max_column
        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Guess the colour"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #appending scores
        if score>=8:
            c1.value=10
            c2.value=0
        elif score>=5:
            c1.value=5
            c2.value=5
        else:
            c1.value=0
            c2.value=10
    

    else:
        row=sheet.max_row
        col=sheet.max_column

        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Guess the colour"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        
        c1=sheet.cell(row=4,column=col+1)
        c2=sheet.cell(row=5,column=col+1)
    
        #   appending scores
        if score>=8:
            c1.value=10
            c2.value=0
        elif score>=5:
            c1.value=5
            c2.value=5
        else:
            c1.value=0
            c2.value=10

    x=1
    for i in range(1,col+1):
        y=sheet.cell(row=1,column=i)
        if str(y.value)=="Guess the colour":
            z=sheet.cell(row=3,column=i)
            x=int(z.value)+1
        else:
            continue

    change_3=sheet.cell(row=3,column=col+1)
    change_3.value=x  
    

    wb.save('final.xlsx')


    
    str1=input("Do you want to play again? (Yes/No)")
    if(str1=='Yes'):
        continue
    else:
        break
print("Thank you for playing")
